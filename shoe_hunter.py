"""
Daily Pronation Shoe Hunter — Cloud Edition
-------------------------------------------
Runs in GitHub Actions every morning. Searches Google Shopping for specific
stability/pronation running shoes in size 11.5 Wide (2E) under a configurable
price cap, then emails a summary and commits results to the repo.

Local test:
    python shoe_hunter.py --no-email
"""

from __future__ import annotations

import argparse
import datetime as dt
import logging
import os
import re
import smtplib
import sys
import time
from dataclasses import dataclass
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders
from pathlib import Path
from urllib.parse import quote_plus

import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

# ---------------------------------------------------------------------------
# CONFIG — shoes to hunt
# ---------------------------------------------------------------------------

TARGET_SHOES = [
    ("Brooks", "Adrenaline GTS", "Brooks Adrenaline GTS"),
    ("Brooks", "Ghost",          "Brooks Ghost"),
    ("Brooks", "Beast",          "Brooks Beast"),
    ("Brooks", "Glycerin GTS",   "Brooks Glycerin GTS"),
    ("Hoka",   "Gaviota",        "Hoka Gaviota"),
    ("Hoka",   "Arahi",          "Hoka Arahi"),
    ("Asics",  "Gel-Kayano",     "Asics Gel-Kayano"),
]

SIZE_TOKENS = ["11.5 wide", "11.5w", "11.5 2e", "size 11.5 wide"]

PRICE_CAP            = 120.00  # max price to flag
GOOD_DEAL_THRESHOLD  = 100.00  # green highlight in Excel
REQUEST_TIMEOUT      = 15
POLITENESS_DELAY     = 2.5

USER_AGENT = (
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
    "AppleWebKit/537.36 (KHTML, like Gecko) "
    "Chrome/124.0 Safari/537.36"
)

# Email config — pulled from environment (GitHub Secrets in cloud, .env locally)
EMAIL_CONFIG = {
    "smtp_server": os.environ.get("SMTP_SERVER", "smtp.gmail.com"),
    "smtp_port":   int(os.environ.get("SMTP_PORT", "587")),
    "from_addr":   os.environ.get("EMAIL_FROM", ""),
    "to_addr":     os.environ.get("EMAIL_TO",   ""),
    "username":    os.environ.get("EMAIL_USER", ""),
    "password":    os.environ.get("EMAIL_PASS", ""),
}

OUTPUT_DIR = Path(__file__).parent / "results"
LOG_DIR    = Path(__file__).parent / "logs"

# ---------------------------------------------------------------------------
# LOGGING
# ---------------------------------------------------------------------------

LOG_DIR.mkdir(exist_ok=True)
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s  %(levelname)-7s  %(message)s",
    handlers=[
        logging.FileHandler(LOG_DIR / f"hunter_{dt.date.today().isoformat()}.log"),
        logging.StreamHandler(sys.stdout),
    ],
)
log = logging.getLogger("shoe_hunter")

# ---------------------------------------------------------------------------
# DATA MODEL
# ---------------------------------------------------------------------------

@dataclass
class ShoeListing:
    brand: str
    model: str
    title: str
    price: float
    retailer: str
    url: str
    size_mentioned: bool
    found_at: str

# ---------------------------------------------------------------------------
# SEARCHER
# ---------------------------------------------------------------------------

class GoogleShoppingSearcher:
    BASE = "https://www.google.com/search"

    def __init__(self):
        self.session = requests.Session()
        self.session.headers.update({
            "User-Agent": USER_AGENT,
            "Accept-Language": "en-US,en;q=0.9",
        })

    def search(self, query: str) -> list[dict]:
        params = {"q": query, "tbm": "shop", "hl": "en", "gl": "us", "num": "40"}
        url = f"{self.BASE}?{'&'.join(f'{k}={quote_plus(v)}' for k,v in params.items())}"
        log.info(f"  query: {query}")
        try:
            resp = self.session.get(url, timeout=REQUEST_TIMEOUT)
            resp.raise_for_status()
        except requests.RequestException as e:
            log.warning(f"  request failed: {e}")
            return []
        return self._parse(resp.text)

    def _parse(self, html: str) -> list[dict]:
        soup = BeautifulSoup(html, "html.parser")
        results: list[dict] = []
        candidates = soup.select(
            ".sh-dgr__content, .sh-dlr__list-result, .sh-pr__product-results, "
            ".KZmu8e, [data-docid], .mnIHsc"
        )
        if not candidates:
            candidates = [el for el in soup.find_all(["div", "li"])
                          if el.find(string=re.compile(r"\$\s?\d+"))]

        for card in candidates:
            text  = card.get_text(" ", strip=True)
            price = self._extract_price(text)
            if price is None:
                continue
            link_el = card.find("a", href=True)
            url = link_el["href"] if link_el else ""
            if url.startswith("/url?"):
                m = re.search(r"[?&]url=([^&]+)", url)
                if m:
                    url = requests.utils.unquote(m.group(1))
            if url.startswith("/"):
                url = "https://www.google.com" + url
            title_el = card.find(["h3", "h4"]) or link_el
            title = title_el.get_text(" ", strip=True) if title_el else text[:80]
            retailer = self._extract_retailer(card, url)
            results.append({
                "title":    title,
                "price":    price,
                "url":      url,
                "retailer": retailer,
                "raw_text": text[:300],
            })
        return results

    @staticmethod
    def _extract_price(text: str) -> float | None:
        m = re.search(r"\$\s?(\d{1,3}(?:,\d{3})*(?:\.\d{1,2})?)", text)
        if not m:
            return None
        try:
            return float(m.group(1).replace(",", ""))
        except ValueError:
            return None

    @staticmethod
    def _extract_retailer(card, url: str) -> str:
        for sel in [".aULzUe", ".IuHnof", ".E5ocAb", "cite"]:
            el = card.select_one(sel)
            if el and el.get_text(strip=True):
                return el.get_text(strip=True)
        m = re.search(r"https?://(?:www\.)?([^/]+)", url)
        return m.group(1) if m else "unknown"

# ---------------------------------------------------------------------------
# HUNT
# ---------------------------------------------------------------------------

def is_size_mentioned(text: str) -> bool:
    t = text.lower()
    return any(tok in t for tok in SIZE_TOKENS)

def hunt(price_cap: float) -> list[ShoeListing]:
    searcher = GoogleShoppingSearcher()
    hits: list[ShoeListing] = []
    seen_urls: set[str] = set()

    for brand, model, name in TARGET_SHOES:
        log.info(f"Searching: {name}")
        for q in (f"{name} mens 11.5 wide", f"{name} mens size 11.5 2E"):
            results = searcher.search(q)
            for r in results:
                if r["url"] in seen_urls:
                    continue
                if r["price"] > price_cap:
                    continue
                first_word = model.lower().split()[0]
                if first_word not in r["title"].lower() \
                   and first_word not in r["raw_text"].lower():
                    continue
                seen_urls.add(r["url"])
                hits.append(ShoeListing(
                    brand=brand,
                    model=model,
                    title=r["title"][:150],
                    price=r["price"],
                    retailer=r["retailer"],
                    url=r["url"],
                    size_mentioned=is_size_mentioned(r["title"] + " " + r["raw_text"]),
                    found_at=dt.datetime.now().isoformat(timespec="seconds"),
                ))
            time.sleep(POLITENESS_DELAY)

    hits.sort(key=lambda h: h.price)
    return hits

# ---------------------------------------------------------------------------
# OUTPUT
# ---------------------------------------------------------------------------

def write_excel(hits: list[ShoeListing], out_path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Hits"
    headers = ["Brand", "Model", "Price", "Retailer",
               "Size 11.5W Mentioned?", "Title", "URL", "Found At"]
    ws.append(headers)

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1F4E78")
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    good_fill   = PatternFill("solid", fgColor="C6EFCE")
    decent_fill = PatternFill("solid", fgColor="FFEB9C")

    for h in hits:
        ws.append([h.brand, h.model, h.price, h.retailer,
                   "Yes" if h.size_mentioned else "Verify",
                   h.title, h.url, h.found_at])
        row_idx = ws.max_row
        fill = good_fill if h.price <= GOOD_DEAL_THRESHOLD else decent_fill
        for cell in ws[row_idx]:
            cell.fill = fill
        ws.cell(row=row_idx, column=3).number_format = '"$"#,##0.00'

    widths = [10, 22, 10, 22, 22, 50, 60, 20]
    for col, w in enumerate(widths, start=1):
        ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = w
    ws.freeze_panes = "A2"
    wb.save(out_path)
    log.info(f"wrote {len(hits)} hits to {out_path}")

def build_email_body(hits: list[ShoeListing]) -> str:
    if not hits:
        return ("No hits today under the price cap. The hunt re-runs tomorrow.\n\n"
                "(Stability shoes in 11.5 Wide under $120 are rare — best windows "
                "are end-of-model-year sales in Aug/Sep and Black Friday.)")
    lines = [f"Found {len(hits)} listing(s) under cap. Cheapest first:\n"]
    for h in hits[:25]:
        flag = "  [VERIFY SIZE]" if not h.size_mentioned else ""
        lines.append(
            f"  ${h.price:>6.2f}  {h.brand:<7} {h.model:<18} @ {h.retailer}{flag}\n"
            f"            {h.url}\n"
        )
    if len(hits) > 25:
        lines.append(f"\n...and {len(hits)-25} more in the attached Excel.")
    return "\n".join(lines)

def send_email(subject: str, body: str, attachment: Path | None) -> bool:
    cfg = EMAIL_CONFIG
    if not all([cfg["from_addr"], cfg["to_addr"], cfg["username"], cfg["password"]]):
        log.info("email credentials not set — skipping send")
        return False

    msg = MIMEMultipart()
    msg["From"]    = cfg["from_addr"]
    msg["To"]      = cfg["to_addr"]
    msg["Subject"] = subject
    msg.attach(MIMEText(body, "plain"))

    if attachment and attachment.exists():
        with open(attachment, "rb") as f:
            part = MIMEBase("application", "octet-stream")
            part.set_payload(f.read())
        encoders.encode_base64(part)
        part.add_header("Content-Disposition",
                        f'attachment; filename="{attachment.name}"')
        msg.attach(part)

    try:
        with smtplib.SMTP(cfg["smtp_server"], cfg["smtp_port"]) as s:
            s.starttls()
            s.login(cfg["username"], cfg["password"])
            s.send_message(msg)
        log.info(f"email sent to {cfg['to_addr']}")
        return True
    except Exception as e:
        log.error(f"email failed: {e}")
        return False

# ---------------------------------------------------------------------------
# ENTRY POINT
# ---------------------------------------------------------------------------

def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--price-cap", type=float, default=PRICE_CAP)
    ap.add_argument("--no-email", action="store_true")
    args = ap.parse_args()

    OUTPUT_DIR.mkdir(exist_ok=True)
    log.info(f"=== Shoe Hunter run @ {dt.datetime.now()} ===")
    log.info(f"Price cap: ${args.price_cap:.2f}   "
             f"Good-deal threshold: ${GOOD_DEAL_THRESHOLD:.2f}")

    hits = hunt(price_cap=args.price_cap)

    today = dt.date.today().isoformat()
    xlsx  = OUTPUT_DIR / f"shoe_hits_{today}.xlsx"
    write_excel(hits, xlsx)

    body = build_email_body(hits)
    subj = f"Shoe Hunter — {len(hits)} hit(s) on {today}"
    print("\n" + "=" * 60)
    print(subj)
    print("=" * 60)
    print(body)
    print("=" * 60)

    if not args.no_email:
        send_email(subj, body, xlsx)

    log.info("=== run complete ===\n")

if __name__ == "__main__":
    main()
